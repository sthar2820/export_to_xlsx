
#code_1
# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# from io import BytesIO
# from datetime import datetime
# import re
# from dateutil import parser


# KNOWN_PLANTS = {
#     "Bielsko Biala", "Birmingham", "Blatna", "Einbeck", "Forsheda",
#     "Olofstrom", "Rotenburg", "Celaya", "Dickson", "Goshen",
#     "Kalamazoo", "Saltillo", "Valley City", "Wellington"
# }


# # def detect_metric_columns(sheet, stop_at_keywords=None):
# #     if stop_at_keywords is None:
# #         stop_at_keywords = [
# #             "demon-strated rate at 100%", "demonstrated rate at 100%",
# #             "demon-strated rate", "demonstrated rate",
# #         ]

# #     metric_cols = []
# #     headers = {}
# #     stop_column_found = None

# #     try:
# #         for search_row in range(1, min(6, sheet.max_row + 1)):
# #             temp_cols = []
# #             temp_headers = {}
# #             temp_stop_col = None

# #             for col in range(3, min(sheet.max_column + 1, 20)):
# #                 try:
# #                     cell = sheet.cell(row=search_row, column=col).value
# #                     if cell and isinstance(cell, str) and len(cell.strip()) > 1:
# #                         header_clean = ' '.join(str(cell).strip().split()).lower()
# #                         temp_headers[col] = header_clean
# #                         temp_cols.append(col)

# #                         for stop_keyword in stop_at_keywords:
# #                             if stop_keyword in header_clean:
# #                                 temp_stop_col = stop_keyword
# #                                 break
# #                         if temp_stop_col:
# #                             break
# #                 except:
# #                     continue

# #             if temp_stop_col or len(temp_headers) > len(headers):
# #                 headers = temp_headers
# #                 metric_cols = temp_cols
# #                 if temp_stop_col:
# #                     stop_column_found = temp_stop_col
# #                     break

# #         if not metric_cols:
# #             metric_cols = list(range(3, min(8, sheet.max_column + 1)))
# #             for col in metric_cols:
# #                 headers[col] = f"column_{chr(64 + col)}".lower()

# #     except:
# #         metric_cols = [3, 4, 5, 6]
# #         headers = {3: "column_c", 4: "column_d", 5: "column_e", 6: "column_f"}

# #     return metric_cols, headers, stop_column_found
# def find_last_row_for_delta_column(sheet, headers, target_metric="Δ (plex std. -> act) $ / piece"):
#     target_col = None
#     for col, header in headers.items():
#         if target_metric.lower() in header.lower():
#             target_col = col
#             break

#     if not target_col:
#         return sheet.max_row  # Fallback: don't truncate

#     last_valid_row = None
#     for row in range(1, sheet.max_row + 1):
#         val = sheet.cell(row=row, column=target_col).value
#         if isinstance(val, (int, float)):
#             last_valid_row = row

#     return last_valid_row if last_valid_row else sheet.max_row

# def detect_metric_columns(sheet, stop_at_keywords=None):
#     if stop_at_keywords is None:
#         stop_at_keywords = [
#              "$ / Piece", "quoted cost/hr"
#             # Removed "demonstrated rate" to continue past it and capture delta columns
#         ]

#     metric_cols = []
#     headers = {}
#     stop_column_found = None

#     try:
#         for search_row in range(1, min(6, sheet.max_row + 1)):
#             temp_cols = []
#             temp_headers = {}
#             temp_stop_col = None

#             for col in range(3, min(sheet.max_column + 1, 20)):
#                 try:
#                     cell = sheet.cell(row=search_row, column=col).value
#                     if cell and isinstance(cell, str) and len(cell.strip()) > 1:
#                         header_clean = ' '.join(str(cell).strip().split()).lower()
#                         temp_headers[col] = header_clean
#                         temp_cols.append(col)

#                         for stop_keyword in stop_at_keywords:
#                             if stop_keyword in header_clean:
#                                 temp_stop_col = stop_keyword
#                                 break
#                         if temp_stop_col:
#                             break
#                 except:
#                     continue

#             if temp_stop_col or len(temp_headers) > len(headers):
#                 headers = temp_headers
#                 metric_cols = temp_cols
#                 if temp_stop_col:
#                     stop_column_found = temp_stop_col
#                     break

#         if not metric_cols:
#             metric_cols = list(range(3, min(12, sheet.max_column + 1)))  # Extended to column 12 to capture more columns
#             for col in metric_cols:
#                 headers[col] = f"column_{chr(64 + col)}".lower()

#     except:
#         metric_cols = [3, 4, 5, 6, 7, 8, 9, 10, 11]  # Extended default range
#         headers = {3: "column_c", 4: "column_d", 5: "column_e", 6: "column_f", 
#                   7: "column_g", 8: "column_h", 9: "column_i", 10: "column_j", 11: "column_k"}

#     return metric_cols, headers, stop_column_found

# def detect_categories(sheet):
#     categories = []
#     category_map = {
#         'S': 'Sales Price', 'M': 'Material', 'I': 'Investment',
#         'T': 'Tooling', 'C': 'Cycle Times', 'H': 'Headcount'
#     }

#     try:
#         for col in range(1, min(4, sheet.max_column + 1)):
#             for row in range(1, min(sheet.max_row + 1, 50)):
#                 try:
#                     val = sheet.cell(row=row, column=col).value
#                     if not val:
#                         continue
#                     text = str(val).strip()
#                     lines = text.split('\n') if '\n' in text else [text]
#                     for line in lines:
#                         line_clean = line.strip().upper()
#                         if len(line_clean) == 1 and line_clean in category_map:
#                             if not any(c['letter'] == line_clean for c in categories):
#                                 categories.append({
#                                     'row': row, 'column': col,
#                                     'letter': line_clean,
#                                     'name': category_map[line_clean]
#                                 })
#                             break
#                 except:
#                     continue
#         categories.sort(key=lambda x: x['row'])
#     except Exception as e:
#         st.error(f"Error detecting categories: {e}")
#         categories = []

#     return categories

# def detect_plant(sheet):
#     for row in range(1, sheet.max_row + 1):
#         for col in range(1, sheet.max_column + 1):
#             val = sheet.cell(row=row, column=col).value
#             if val and isinstance(val, str):
#                 text = val.strip()
#                 for plant in KNOWN_PLANTS:
#                     if plant.lower() in text.lower():
#                         return plant, row
#     return None, None

# def detect_part_name(sheet, categories):
#     try:
#         if not categories:
#             return None
#         first_category_row = categories[0]['row']
#         for row in range(first_category_row - 1, 0, -1):  # search upward
#             val = sheet.cell(row=row, column=2).value  
#             if val and isinstance(val, str):
#                 val = val.strip()
#                 # Return first non-empty text that isn't a single letter
#                 if len(val) > 3 and val.upper() not in {'S', 'M', 'I', 'T', 'C', 'H'}:
#                     return val
#     except:
#         pass
#     return None
      

# def extract_date(text):
#     if not isinstance(text, str):
#         return None

#     matches = re.findall(r"\b\d{1,2}[/-]\d{2,4}(?:[/-]\d{2,4})?\b", text)
#     for match in matches:
#         try:
#             if re.match(r"\d{1,2}/\d{4}$", match):  # MM/YYYY
#                 dt = datetime.strptime(match, "%m/%Y")
#             elif re.match(r"\d{1,2}/\d{2}$", match):  # MM/YY
#                 dt = datetime.strptime(match, "%m/%y")
#             elif re.match(r"\d{1,2}/\d{1,2}/\d{4}$", match):  # MM/DD/YYYY
#                 dt = datetime.strptime(match, "%m/%d/%Y")
#             else:
#                 continue
#             return dt.strftime("%Y-%m-%d")
#         except:
#             continue
#     return None



# def find_subcategory_column(sheet, categories):
#     try:
#         if not categories:
#             return 3
#         category_col = categories[0]['column']
#         candidates = [category_col + 1, category_col + 2, 3, 2]
#         best_col = category_col + 1
#         max_text_cells = 0

#         for col in candidates:
#             if col < 1 or col > sheet.max_column:
#                 continue
#             text_cells = 0
#             for row in range(1, min(30, sheet.max_row + 1)):
#                 cell = sheet.cell(row=row, column=col).value
#                 if cell and isinstance(cell, str) and len(cell.strip()) >= 2:
#                     text_cells += 1
#             if text_cells > max_text_cells:
#                 max_text_cells = text_cells
#                 best_col = col
#         return best_col
#     except:
#         return 3

# def extract_smitch_data(sheet, categories, metric_cols, headers, subcategory_col, plant_name=None, part_name=None):
#     extracted = []

#     if not categories:
#         st.warning("No categories found")
#         return []

#     METRIC_NORMALIZATION = {
#         "quoted cost model": "Quoted",
#         "quoted": "Quoted",
#         "plex standard": "Plex",
#         "plex": "Plex",
#         "actual performance": "Actual",
#         "actual": "Actual",
#         "forecasted cost": "Forecasted",
#         "forecasted": "Forecasted",
#         "demonstrated rate": "Demonstrated",
#         "demon-strated": "Demonstrated",
#         # Delta columns
#         "Δ (quote -> actual) JPH": "Quoted_JPH",
#         "Δ (quote -> act) $ / Piece": "Quoted_$", 
#         "Δ (plex std -> actual) JPH": "Plex_JPH",
#         "Δ (plex std. -> act) $ / Piece": "Plex_$",
#         # "delta quote": "Δ Quote→Actual",
#         # "delta plex": "Δ Plex→Actual",
#         # "quote -> actual": "Δ Quote→Actual",
#         # "plex std -> actual": "Δ Plex→Actual",
#     }

#     col_date_map = {}
#     for col in metric_cols:
#         date_found = None
#         for row in range(1, 6):
#             cell_val = sheet.cell(row=row, column=col).value
#             if isinstance(cell_val, str):
#                 possible_date = extract_date(cell_val)
#                 if possible_date:
#                     date_found = possible_date
#                     break
#         col_date_map[col] = date_found

#     for i in range(len(categories)):
#         current = categories[i]
#         start_row = current['row']
#         if i + 1 < len(categories):
#             end_row = categories[i + 1]['row'] - 1
#         else:
#             end_row = find_last_row_for_delta_column(sheet, headers)



#         for row in range(start_row, end_row + 1):
#             subcat_cell = sheet.cell(row=row, column=subcategory_col).value
#             if not subcat_cell:
#                 continue
#             subcat = str(subcat_cell).strip()

#             for col in metric_cols:
#                 val = sheet.cell(row=row, column=col).value
#                 if val is None:
#                     continue

#                 # Handle different data types (numbers, formulas, text with numbers)
#                 if isinstance(val, (int, float)):
#                     numeric_value = float(val)
#                 else:
#                     # Try to extract number from text/formula
#                     val_str = str(val).strip()
#                     if not val_str:
#                         continue
#                     try:
#                         # Look for numbers in the string (handles formulas, text with numbers)
#                         import re
#                         numeric_matches = re.findall(r"[-+]?\d*\.?\d+", val_str)
#                         if not numeric_matches:
#                             continue
#                         numeric_value = float(numeric_matches[0])
#                     except (ValueError, IndexError):
#                         continue

#                 raw_header = headers.get(col, f"Column_{chr(64 + col)}").strip().lower().split('\n')[0]
#                 matched_key = next((k for k in METRIC_NORMALIZATION if k in raw_header), None)
 

#                 if matched_key:
#                     metric = METRIC_NORMALIZATION[matched_key]
#                 else:
#                     metric = raw_header.split()[0].capitalize() if raw_header else f"Col_{col}"

#                 date_str = col_date_map.get(col)

#                 entry = {
#                     'Category': current['name'],
#                     'Subcategory': subcat,
#                     'Date': date_str,
#                     'Metric': metric,
#                     'Value': numeric_value
#                 }
#                 if plant_name:
#                     entry['Plant'] = plant_name
#                 if part_name:
#                     entry['Part Name'] = part_name

#                 extracted.append(entry)

#     return extracted
# # def extract_smitch_data(sheet, categories, metric_cols, headers, subcategory_col, plant_name=None, part_name=None):
# #     extracted = []

# #     if not categories:
# #         st.warning("No categories found")
# #         return []

# #     METRIC_NORMALIZATION = {
# #         "quoted cost model": "Quoted",
# #         "quoted": "Quoted",
# #         "plex standard": "Plex",
# #         "plex": "Plex",
# #         "actual performance": "Actual",
# #         "actual": "Actual",
# #         "forecasted cost": "Forecasted",
# #         "forecasted": "Forecasted",
# #         "demonstrated rate": "Demonstrated",
# #         "demon-strated": "Demonstrated",
# #     }

# #     # Preprocess: Extract date from headers for each metric column
# #     col_date_map = {}
# #     for col in metric_cols:
# #         date_found = None
# #         for row in range(1, 6):
# #             cell_val = sheet.cell(row=row, column=col).value
# #             if isinstance(cell_val, str):
# #                 possible_date = extract_date(cell_val)
# #                 if possible_date:
# #                     date_found = possible_date
# #                     break
# #         col_date_map[col] = date_found

# #     # Iterate through category rows
# #     for i in range(len(categories)):
# #         current = categories[i]
# #         start_row = current['row']
# #         end_row = categories[i + 1]['row'] - 1 if i + 1 < len(categories) else min(start_row + 25, sheet.max_row)

# #         for row in range(start_row, end_row + 1):
# #             subcat_cell = sheet.cell(row=row, column=subcategory_col).value
# #             if not subcat_cell:
# #                 continue
# #             subcat = str(subcat_cell).strip()

# #             for col in metric_cols:
# #                 val = sheet.cell(row=row, column=col).value
# #                 if not isinstance(val, (int, float)):
# #                     continue

# #                 # Normalize header
# #                 raw_header = headers.get(col, f"column_{chr(64 + col)}").strip().lower().split('\n')[0]
# #                 if "cm%" in raw_header:
# #                     continue

# #                 # Clean header for matching
# #                 cleaned_header = re.sub(r'[^a-z\s$\/→-]', '', raw_header.lower()).strip()

# #                 # Match against normalization dict
# #                 matched_key = next((k for k in METRIC_NORMALIZATION if k in cleaned_header), None)

# #                 if matched_key:
# #                     metric = METRIC_NORMALIZATION[matched_key]
# #                 elif "quoted jph" in cleaned_header:
# #                     metric = "Quoted_JPH"
# #                 elif "quoted $" in cleaned_header or "quoted $ / piece" in cleaned_header:
# #                     metric = "Quoted_$"
# #                 elif "actual jph" in cleaned_header:
# #                     metric = "Actual_JPH"
# #                 elif "actual $" in cleaned_header or "actual $ / piece" in cleaned_header:
# #                     metric = "Actual_$"
# #                 elif "plex std" in cleaned_header and "jph" in cleaned_header:
# #                     metric = "Plex_JPH"
# #                 elif "plex std" in cleaned_header and ("$" in cleaned_header or "piece" in cleaned_header):
# #                     metric = "Plex_$"
# #                 else:
# #                     metric = raw_header.split()[0].capitalize() if raw_header else f"Col_{col}"

# #                 date_str = col_date_map.get(col)

# #                 entry = {
# #                     'Category': current['name'],
# #                     'Subcategory': subcat,
# #                     'Date': date_str,
# #                     'Metric': metric,
# #                     'Value': float(val)
# #                 }
# #                 if plant_name:
# #                     entry['Plant'] = plant_name
# #                 if part_name:
# #                     entry['Part Name'] = part_name

# #                 extracted.append(entry)

# #     return extracted


# st.title("📊 SMITCH Excel Extractor")
# st.write("Upload SMITCH Excel files to extract structured data")

# uploaded_files = st.file_uploader("Choose Excel files", type=["xlsm", "xlsx"], accept_multiple_files=True)

# if uploaded_files:
#     st.write(f"Processing {len(uploaded_files)} file(s)...")
#     for file in uploaded_files:
#         st.subheader(f"{file.name}")
#         try:
#             # wb = load_workbook(file, data_only=True)
#             wb = load_workbook(filename=BytesIO(file.getvalue()), data_only=True)

#             ws = wb.active
#             st.write(f"File loaded: {ws.max_row} rows × {ws.max_column} columns")
#             with st.spinner("Detecting file structure..."):
#                 metric_columns, headers, stop_column_found = detect_metric_columns(ws)
#                 category_rows = detect_categories(ws)
#                 subcategory_col = find_subcategory_column(ws, category_rows)
#                 plant_name, plant_row = detect_plant(ws)
#                 smitch_row = category_rows[0]['row'] if category_rows else ws.max_row
#                 part_name = detect_part_name(ws, category_rows)

#             col1, col2, col3, col4 = st.columns(4)
#             with col1:
#                 st.metric("Categories", len(category_rows))
#             with col2:
#                 st.metric("Metric Columns", len(metric_columns))
#             with col3:
#                 st.metric("Subcategory Col", chr(64 + subcategory_col))
#             with col4:
#                 st.metric("Stop Column", stop_column_found.title() if stop_column_found else "Auto-detected")

#             with st.spinner("Extracting data..."):
#                 data = extract_smitch_data(ws, category_rows, metric_columns, headers, subcategory_col, plant_name, part_name)

#             if data:
#                 df = pd.DataFrame(data)
#                 st.success(f"Extracted {len(df)} records")

#                 st.write("**Categories found:**")
#                 for cat, count in df['Category'].value_counts().items():
#                     st.write(f"• {cat}: {count} records")

#                 st.write("**Data preview:**")
#                 st.dataframe(df.head(10))

#                 output = BytesIO()
#                 with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
#                     df.to_excel(writer, index=False, sheet_name='Extracted')
#                 output.seek(0)

#                 st.download_button(
#                     label=" Download Excel",
#                     data=output,
#                     file_name=f"{file.name.split('.')[0]}_extracted.xlsx",
#                     mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#                 )
#             else:
#                 st.warning(" No data extracted from this file")
#         except Exception as e:
#             st.error(f" Failed to process {file.name}")
#             st.error(f"Error: {str(e)}")
# else:
#     st.info(" Upload Excel files to get started")

##code_2
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
import re
from dateutil import parser


KNOWN_PLANTS = {
    "Bielsko Biala", "Birmingham", "Blatna", "Einbeck", "Forsheda",
    "Olofstrom", "Rotenburg", "Celaya", "Dickson", "Goshen",
    "Kalamazoo", "Saltillo", "Valley City", "Wellington"
}

def extract_weekly_apw(sheet, plant_name=None, part_name=None):
    for row in range(1, 21):  
        for col in range(1, 31): 
            cell = sheet.cell(row=row, column=col).value
            if isinstance(cell, str) and "WEEKLY APW" in cell.upper():
                # Look for numeric value to the right
                for value_col in range(col + 1, min(col + 10, sheet.max_column + 1)):
                    val = sheet.cell(row=row, column=value_col).value
                    numeric_value = extract_numeric_value(val)
                    if numeric_value is not None:
                        metric_name = find_apw_metric_context(sheet, row, value_col)
                        return [{
                            "Category": "EBIT LOSS",
                            "Subcategory": "Weekly APW",
                            "Metric": metric_name,
                            "Value": numeric_value,
                            "Plant": plant_name,
                            "Part Name": part_name
                        }]
    return []

def extract_numeric_value(val):
    if isinstance(val, (int, float)):
        return float(val)
    elif isinstance(val, str):
        clean = re.sub(r'[^\d.\-]', '', val)
        try:
            return float(clean)
        except:
            return None
    return None

def find_apw_metric_context(sheet, row, col):
    offsets = [(-1, 0), (-2, 0), (0, -1), (0, 1)]
    for dr, dc in offsets:
        r, c = row + dr, col + dc
        if r >= 1 and c >= 1:
            val = sheet.cell(row=r, column=c).value
            if isinstance(val, str) and len(val.strip()) > 2:
                text = val.strip().replace('\n', ' ').title()
                if any(word in text.lower() for word in ["quoted", "plex", "actual"]):
                    return text
    return "Weekly APW Value"

def detect_metric_columns(sheet, stop_at_keywords=None):
    if stop_at_keywords is None:
        stop_at_keywords = [
            "demon-strated rate at 100%", "demonstrated rate at 100%",
            "demon-strated rate", "demonstrated rate",
        ]

    metric_cols = []
    headers = {}
    stop_column_found = None

    try:
        for search_row in range(1, min(6, sheet.max_row + 1)):
            temp_cols = []
            temp_headers = {}
            temp_stop_col = None

            for col in range(3, min(sheet.max_column + 1, 20)):
                try:
                    cell = sheet.cell(row=search_row, column=col).value
                    if cell and isinstance(cell, str) and len(cell.strip()) > 1:
                        header_clean = ' '.join(str(cell).strip().split()).lower()
                        temp_headers[col] = header_clean
                        temp_cols.append(col)

                        for stop_keyword in stop_at_keywords:
                            if stop_keyword in header_clean:
                                temp_stop_col = stop_keyword
                                break
                        if temp_stop_col:
                            break
                except:
                    continue

            if temp_stop_col or len(temp_headers) > len(headers):
                headers = temp_headers
                metric_cols = temp_cols
                if temp_stop_col:
                    stop_column_found = temp_stop_col
                    break

        if not metric_cols:
            metric_cols = list(range(3, min(8, sheet.max_column + 1)))
            for col in metric_cols:
                headers[col] = f"column_{chr(64 + col)}".lower()

    except:
        metric_cols = [3, 4, 5, 6]
        headers = {3: "column_c", 4: "column_d", 5: "column_e", 6: "column_f"}

    return metric_cols, headers, stop_column_found


def detect_categories(sheet):
    categories = []
    category_map = {
        'S': 'Sales Price', 'M': 'Material', 'I': 'Investment',
        'T': 'Tooling', 'C': 'Cycle Times', 'H': 'Headcount'
    }

    try:
        for col in range(1, min(4, sheet.max_column + 1)):
            for row in range(1, min(sheet.max_row + 1, 50)):
                try:
                    val = sheet.cell(row=row, column=col).value
                    if not val:
                        continue
                    text = str(val).strip()
                    lines = text.split('\n') if '\n' in text else [text]
                    for line in lines:
                        line_clean = line.strip().upper()
                        if len(line_clean) == 1 and line_clean in category_map:
                            if not any(c['letter'] == line_clean for c in categories):
                                categories.append({
                                    'row': row, 'column': col,
                                    'letter': line_clean,
                                    'name': category_map[line_clean]
                                })
                            break
                except:
                    continue
        categories.sort(key=lambda x: x['row'])
    except Exception as e:
        st.error(f"Error detecting categories: {e}")
        categories = []

    return categories

def detect_plant(sheet):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row, column=col).value
            if val and isinstance(val, str):
                text = val.strip()
                for plant in KNOWN_PLANTS:
                    if plant.lower() in text.lower():
                        return plant, row
    return None, None

def detect_part_name(sheet, categories):
    try:
        if not categories:
            return None
        first_category_row = categories[0]['row']
        for row in range(first_category_row - 1, 0, -1):  # search upward
            val = sheet.cell(row=row, column=2).value  
            if val and isinstance(val, str):
                val = val.strip()
                # Return first non-empty text that isn't a single letter
                if len(val) > 3 and val.upper() not in {'S', 'M', 'I', 'T', 'C', 'H'}:
                    return val
    except:
        pass
    return None
      

def extract_date(text):
    if not isinstance(text, str):
        return None

    matches = re.findall(r"\b\d{1,2}[/-]\d{2,4}(?:[/-]\d{2,4})?\b", text)
    for match in matches:
        try:
            if re.match(r"\d{1,2}/\d{4}$", match):  # MM/YYYY
                dt = datetime.strptime(match, "%m/%Y")
            elif re.match(r"\d{1,2}/\d{2}$", match):  # MM/YY
                dt = datetime.strptime(match, "%m/%y")
            elif re.match(r"\d{1,2}/\d{1,2}/\d{4}$", match):  # MM/DD/YYYY
                dt = datetime.strptime(match, "%m/%d/%Y")
            else:
                continue
            return dt.strftime("%Y-%m-%d")
        except:
            continue
    return None



def find_subcategory_column(sheet, categories):
    try:
        if not categories:
            return 3
        category_col = categories[0]['column']
        candidates = [category_col + 1, category_col + 2, 3, 2]
        best_col = category_col + 1
        max_text_cells = 0

        for col in candidates:
            if col < 1 or col > sheet.max_column:
                continue
            text_cells = 0
            for row in range(1, min(30, sheet.max_row + 1)):
                cell = sheet.cell(row=row, column=col).value
                if cell and isinstance(cell, str) and len(cell.strip()) >= 2:
                    text_cells += 1
            if text_cells > max_text_cells:
                max_text_cells = text_cells
                best_col = col
        return best_col
    except:
        return 3



# def extract_oh_metrics(sheet, plant_name=None, part_name=None, categories=None):
#     """
#     Extract EBIT metrics for OH subcategories only
#     """
#     extracted = []
#     metric_map = {
#         "quoted cost/pc": "Quoted_Cost",
#         "actual oee cost/pc at plex cost/hr (quote)": "Actual_OEE", 
#         "plex standard cost/pc": "Plex_Cost",
#         "actual oee cost/pc at plex cost/hr (plex)": "Plex_OEE"
#     }
#     allowed_metrics = set(metric_map.values())
    
#     for row in range(1, min(sheet.max_row + 1, 100)):
#         for col in range(1, min(sheet.max_column + 1, 30)):
#             val = sheet.cell(row=row, column=col).value
#             if not isinstance(val, str):
#                 continue
            
#             val_clean = val.strip().upper().replace(" ", "")
            
#             # Look for OH
#             if "OH" in val_clean:
#                 subcategory = "OH"
                
#                 category = get_category_from_main(categories, row) if categories else "Unknown"
#                 seen_metrics = set()
                
#                 for c in range(col + 1, min(col + 15, sheet.max_column + 1)):
#                     raw_val = sheet.cell(row=row, column=c).value
#                     if raw_val is None:
#                         continue
                        
#                     try:
#                         if isinstance(raw_val, (int, float)):
#                             value = float(raw_val)
#                         else:
#                             clean_val = str(raw_val).strip().replace("$", "").replace("€", "").replace("£", "").replace(",", "")
#                             if clean_val:
#                                 value = float(clean_val)
#                             else:
#                                 continue
#                     except:
#                         continue

#                     # Search for header
#                     metric = None
#                     for rh in range(row - 1, max(0, row - 10), -1):
#                         header = sheet.cell(row=rh, column=c).value
#                         if isinstance(header, str):
#                             header_lower = header.strip().lower()
#                             for k, v in metric_map.items():
#                                 if k in header_lower:
#                                     metric = v
#                                     break
#                             if metric:
#                                 break

#                     if metric and metric in allowed_metrics and metric not in seen_metrics:
#                         extracted.append({
#                             "Category": category,
#                             "Subcategory": subcategory,
#                             "Metric": metric,
#                             "Value": value,
#                             "Plant": plant_name,
#                             "Part Name": part_name
#                         })
#                         seen_metrics.add(metric)
    
#     return extracted

# def extract_lab_metrics(sheet, plant_name=None, part_name=None, categories=None):
#     """
#     Extract EBIT metrics for LAB subcategories only - FIXED VERSION
#     """
#     extracted = []
#     metric_map = {
#         "quoted cost/pc": "Quoted_Cost",
#         "actual oee cost/pc at plex cost/hr (quote)": "Actual_OEE", 
#         "plex standard cost/pc": "Plex_Cost",
#         "actual oee cost/pc at plex cost/hr (plex)": "Plex_OEE"
#     }
#     allowed_metrics = set(metric_map.values())
    
#     for row in range(1, min(sheet.max_row + 1, 100)):
#         for col in range(1, min(sheet.max_column + 1, 30)):
#             val = sheet.cell(row=row, column=col).value
#             if not isinstance(val, str):
#                 continue
            
#             # FIXED: More flexible LAB detection
#             val_clean = val.strip().upper().replace(" ", "")
            
#             # Look for LAB in any form
#             if "LAB" in val_clean:
#                 subcategory = "LAB"
                
#                 category = get_category_from_main(categories, row) if categories else "Unknown"
#                 seen_metrics = set()
                
#                 for c in range(col + 1, min(col + 15, sheet.max_column + 1)):
#                     raw_val = sheet.cell(row=row, column=c).value
#                     if raw_val is None:
#                         continue
                        
#                     try:
#                         if isinstance(raw_val, (int, float)):
#                             value = float(raw_val)
#                         else:
#                             clean_val = str(raw_val).strip().replace("$", "").replace("€", "").replace("£", "").replace(",", "")
#                             if clean_val:
#                                 value = float(clean_val)
#                             else:
#                                 continue
#                     except:
#                         continue

#                     # Search for header
#                     metric = None
#                     for rh in range(row - 1, max(0, row - 10), -1):
#                         header = sheet.cell(row=rh, column=c).value
#                         if isinstance(header, str):
#                             header_lower = header.strip().lower()
#                             for k, v in metric_map.items():
#                                 if k in header_lower:
#                                     metric = v
#                                     break
#                             if metric:
#                                 break

#                     if metric and metric in allowed_metrics and metric not in seen_metrics:
#                         extracted.append({
#                             "Category": category,
#                             "Subcategory": subcategory,
#                             "Metric": metric,
#                             "Value": value,
#                             "Plant": plant_name,
#                             "Part Name": part_name
#                         })
#                         seen_metrics.add(metric)
    
#     return extracted

# def extract_ebit_metrics(sheet, plant_name=None, part_name=None, categories=None):
#     oh_data = extract_oh_metrics(sheet, plant_name, part_name, categories)
#     lab_data = extract_lab_metrics(sheet, plant_name, part_name, categories)
    
#     return oh_data + lab_data

def extract_ebit_metrics(sheet, plant_name=None, part_name=None, categories=None):
    """
    Extract normalized EBIT metrics for OH and LAB subcategories from Excel sheet.
    Supports both OH and LAB in the same column.
    """
    extracted = []
    metric_map = {
        "quoted cost/pc": "Quoted_Cost",
        "actual oee cost/pc at plex cost/hr (quote)": "Actual_OEE", 
        "plex standard cost/pc": "Plex_Cost",
        "actual oee cost/pc at plex cost/hr (plex)": "Plex_OEE"
    }
    allowed_metrics = set(metric_map.values())

    for row in range(1, min(sheet.max_row + 1, 100)):  # Limit rows for performance
        for col in range(1, min(sheet.max_column + 1, 30)):
            val = sheet.cell(row=row, column=col).value
            if not isinstance(val, str):
                continue

            val_upper = val.strip().upper()
            subcategories_found = []
            if "OH" in val_upper and len(val_upper) <= 25:
                subcategories_found = "OH"
            if "LAB" in val_upper and len(val_upper) <= 25:
                subcategories_found = "LAB"

            if not subcategories_found:
                continue

            category = get_category_from_main(categories, row) if categories else "Unknown"
            seen_metrics = set()

            for c in range(col + 1, min(col + 15, sheet.max_column + 1)):
                raw_val = sheet.cell(row=row, column=c).value
                if raw_val is None:
                    continue

                try:
                    value = float(str(raw_val).strip().replace("$", "").replace(",", ""))
                except:
                    continue

                metric = None
                for rh in range(row - 1, max(0, row - 10), -1):
                    header = sheet.cell(row=rh, column=c).value
                    if isinstance(header, str):
                        header_lower = header.strip().lower()
                        for k, v in metric_map.items():
                            if k in header_lower:
                                metric = v
                                break
                        if metric:
                            break

                if metric and metric in allowed_metrics and metric not in seen_metrics:
                    for subcat in subcategories_found:
                        extracted.append({
                            "Category": category,
                            "Subcategory": subcat,
                            "Metric": metric,
                            "Value": value,
                            "Plant": plant_name,
                            "Part Name": part_name
                        })
                    seen_metrics.add(metric)

    return extracted


def get_category_from_main(categories, target_row):

    if not categories:
        return "Unknown"
    
    # Find the category with the highest row number that's still <= target_row
    best_category = None
    for category in categories:
        if category['row'] <= target_row:
            if best_category is None or category['row'] > best_category['row']:
                best_category = category
    
    return best_category['name'] if best_category else "Unknown"

def extract_smitch_data(sheet, categories, metric_cols, headers, subcategory_col, plant_name=None, part_name=None):
    extracted = []

    if not categories:
        st.warning("No categories found")
        return []

    METRIC_NORMALIZATION = {
        "quoted cost model": "Quoted",
        "quoted": "Quoted",
        "plex standard": "Plex",
        "plex": "Plex",
        "actual performance": "Actual",
        "actual": "Actual",
        "forecasted cost": "Forecasted",
        "forecasted": "Forecasted",
        "demonstrated rate": "Demonstrated",
        "demon-strated": "Demonstrated",
    }

    # Preprocess: Extract date from headers for each metric column
    col_date_map = {}
    for col in metric_cols:
        date_found = None
        for row in range(1, 6):
            cell_val = sheet.cell(row=row, column=col).value
            if isinstance(cell_val, str):
                possible_date = extract_date(cell_val)
                if possible_date:
                    date_found = possible_date
                    break
        col_date_map[col] = date_found

    # Iterate through category rows
    for i in range(len(categories)):
        current = categories[i]
        start_row = current['row']
        end_row = categories[i + 1]['row'] - 1 if i + 1 < len(categories) else min(start_row + 25, sheet.max_row)

        for row in range(start_row, end_row + 1):
            subcat_cell = sheet.cell(row=row, column=subcategory_col).value
            if not subcat_cell:
                continue
            subcat = str(subcat_cell).strip()

            for col in metric_cols:
                val = sheet.cell(row=row, column=col).value
                if not isinstance(val, (int, float)):
                    continue

                # Normalize header
                raw_header = headers.get(col, f"column_{chr(64 + col)}").strip().lower().split('\n')[0]
                if "cm%" in raw_header:
                    continue

                # Clean header for matching
                cleaned_header = re.sub(r'[^a-z\s$\/→-]', '', raw_header.lower()).strip()

                # Match against normalization dict
                matched_key = next((k for k in METRIC_NORMALIZATION if k in cleaned_header), None)

                if matched_key:
                    metric = METRIC_NORMALIZATION[matched_key]
                elif "quoted jph" in cleaned_header:
                    metric = "Quoted_JPH"
                elif "quoted $" in cleaned_header or "quoted $ / piece" in cleaned_header:
                    metric = "Quoted_$"
                elif "actual jph" in cleaned_header:
                    metric = "Actual_JPH"
                elif "actual $" in cleaned_header or "actual $ / piece" in cleaned_header:
                    metric = "Actual_$"
                elif "plex std" in cleaned_header and "jph" in cleaned_header:
                    metric = "Plex_JPH"
                elif "plex std" in cleaned_header and ("$" in cleaned_header or "piece" in cleaned_header):
                    metric = "Plex_$"
                else:
                    metric = raw_header.split()[0].capitalize() if raw_header else f"Col_{col}"

                date_str = col_date_map.get(col)

                entry = {
                    'Category': current['name'],
                    'Subcategory': subcat,
                    'Date': date_str,
                    'Metric': metric,
                    'Value': float(val)
                }
                if plant_name:
                    entry['Plant'] = plant_name
                if part_name:
                    entry['Part Name'] = part_name

                extracted.append(entry)

    return extracted


st.title(" SMITCH Excel Extractor")
st.write("Upload SMITCH Excel files to extract structured data")

uploaded_files = st.file_uploader("Choose Excel files", type=["xlsm", "xlsx"], accept_multiple_files=True)

if uploaded_files:
    st.write(f"Processing {len(uploaded_files)} file(s)...")
    for file in uploaded_files:
        st.subheader(f"{file.name}")
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            st.write(f"File loaded: {ws.max_row} rows × {ws.max_column} columns")
            with st.spinner("Detecting file structure..."):
                metric_columns, headers, stop_column_found = detect_metric_columns(ws)
                category_rows = detect_categories(ws)
                subcategory_col = find_subcategory_column(ws, category_rows)
                plant_name, plant_row = detect_plant(ws)
                smitch_row = category_rows[0]['row'] if category_rows else ws.max_row
                part_name = detect_part_name(ws, category_rows)

            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Categories", len(category_rows))
            with col2:
                st.metric("Metric Columns", len(metric_columns))
            with col3:
                st.metric("Subcategory Col", chr(64 + subcategory_col))
            with col4:
                st.metric("Stop Column", stop_column_found.title() if stop_column_found else "Auto-detected")

            with st.spinner("Extracting data..."):
                data = extract_smitch_data(ws, category_rows, metric_columns, headers, subcategory_col, plant_name, part_name)
                weekly_apw_data = extract_weekly_apw(ws, plant_name, part_name)
                ebit_data = extract_ebit_metrics(ws, plant_name, part_name, category_rows)

                data.extend(weekly_apw_data)
                data.extend(ebit_data)
            if data:
                df = pd.DataFrame(data)
                st.success(f"Extracted {len(df)} records")

                st.write("**Categories found:**")
                for cat, count in df['Category'].value_counts().items():
                    st.write(f"• {cat}: {count} records")

                st.write("**Data preview:**")
                st.dataframe(df.head(10))

                output = BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False, sheet_name='Extracted')
                output.seek(0)

                st.download_button(
                    label=" Download Excel",
                    data=output,
                    file_name=f"{file.name.split('.')[0]}_extracted.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                st.warning(" No data extracted from this file")
        except Exception as e:
            st.error(f" Failed to process {file.name}")
            st.error(f"Error: {str(e)}")
else:
    st.info(" Upload Excel files to get started")

